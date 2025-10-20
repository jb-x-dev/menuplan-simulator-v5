"""
Excel-Export für Menüplan mit ausgewählten Portionen
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_excel_export(plan_data):
    """
    Erstellt einen Excel-Export des Menüplans mit ausgewählten Portionen.
    
    Args:
        plan_data: Dictionary mit Plan-Daten (days, statistics)
    
    Returns:
        BytesIO: Excel-Datei als Byte-Stream
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Menüplan"
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titel
    ws['A1'] = 'Menüplan - Ausgewählte Portionen'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    # Statistiken
    stats = plan_data.get('statistics', {})
    ws['A2'] = f"Zeitraum: {plan_data['days'][0]['date']} bis {plan_data['days'][-1]['date']}"
    ws['A3'] = f"Durchschnitt BKT: {stats.get('average_bkt', 0):.2f}€"
    ws['B3'] = f"Gesamtkosten: {stats.get('total_cost', 0):.2f}€"
    ws['C3'] = f"Tage: {stats.get('total_days', len(plan_data['days']))}"
    
    # Header-Zeile (Zeile 5)
    row = 5
    headers = ['Datum', 'Wochentag', 'Mahlzeit', 'Rezept', 'Portionen', 'Kosten/Portion', 'Gesamtkosten', 'Allergene']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Daten
    row = 6
    for day in plan_data['days']:
        day_date = day['date']
        day_name = datetime.fromisoformat(day_date).strftime('%A')
        
        for menu_line in day['menu_lines']:
            meal_name = menu_line['name']
            
            for recipe_slot in menu_line['recipes']:
                # Prüfe ob es ein MealSlot mit Optionen ist
                if 'options' in recipe_slot and 'selected_index' in recipe_slot:
                    # Neues Format: MealSlot
                    selected_idx = recipe_slot['selected_index']
                    selected_recipe = recipe_slot['options'][selected_idx]
                    
                    recipe_name = selected_recipe['recipe_name']
                    cost_per_serving = selected_recipe['cost_per_serving']
                    portions = recipe_slot.get('target_count', 50)  # Default 50
                    total_cost = cost_per_serving * portions
                    allergens = ', '.join(selected_recipe.get('allergens', []))
                else:
                    # Altes Format: Einzelnes Rezept
                    recipe_name = recipe_slot.get('recipe_name', 'Unbekannt')
                    cost_per_serving = recipe_slot.get('cost_per_serving', 0)
                    portions = recipe_slot.get('target_count', 50)
                    total_cost = cost_per_serving * portions
                    allergens = ', '.join(recipe_slot.get('allergens', []))
                
                # Daten schreiben
                ws.cell(row=row, column=1, value=day_date).border = border
                ws.cell(row=row, column=2, value=day_name).border = border
                ws.cell(row=row, column=3, value=meal_name).border = border
                ws.cell(row=row, column=4, value=recipe_name).border = border
                ws.cell(row=row, column=5, value=portions).border = border
                ws.cell(row=row, column=6, value=f"{cost_per_serving:.2f}€").border = border
                ws.cell(row=row, column=7, value=f"{total_cost:.2f}€").border = border
                ws.cell(row=row, column=8, value=allergens).border = border
                
                row += 1
    
    # Spaltenbreiten anpassen
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    
    # Summenzeile
    row += 1
    ws.cell(row=row, column=1, value='GESAMT').font = subheader_font
    ws.cell(row=row, column=1).fill = subheader_fill
    ws.merge_cells(f'A{row}:F{row}')
    
    # Berechne Gesamtsumme
    total_sum = sum(day['total_cost'] for day in plan_data['days'])
    ws.cell(row=row, column=7, value=f"{total_sum:.2f}€").font = subheader_font
    ws.cell(row=row, column=7).fill = subheader_fill
    
    # Excel-Datei in BytesIO speichern
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

